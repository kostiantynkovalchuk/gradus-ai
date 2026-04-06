"""
HR Admin Dashboard API
Protected by HTTP Basic Auth
"""

from fastapi import APIRouter, Depends, HTTPException, Request, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from sqlalchemy import text
from datetime import date, timedelta
import secrets
import logging
import io
from typing import Optional

from models import get_db
from hunt_config import ROI_CONSTANTS

router = APIRouter(prefix="/hr", tags=["HR Admin"])
security = HTTPBasic()
logger = logging.getLogger(__name__)

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "Maya_2026"


def verify_admin(credentials: HTTPBasicCredentials = Depends(security)):
    """Verify HTTP Basic Auth credentials"""
    correct_username = secrets.compare_digest(credentials.username, ADMIN_USERNAME)
    correct_password = secrets.compare_digest(credentials.password, ADMIN_PASSWORD)
    
    if not (correct_username and correct_password):
        raise HTTPException(
            status_code=401,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username


@router.get("", response_class=HTMLResponse)
async def admin_dashboard(
    request: Request,
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Render admin dashboard HTML"""
    import os
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'hr_admin.html')
    
    try:
        with open(template_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        return HTMLResponse(content=html_content)
    except FileNotFoundError:
        raise HTTPException(status_code=500, detail="Dashboard template not found")


@router.get("/api/overview")
async def get_overview(
    days: int = 7,
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Get overview statistics"""
    db_session = next(get_db())
    
    try:
        result = db_session.execute(text("""
            SELECT
                COUNT(*) as total_queries,
                COUNT(*) FILTER (WHERE preset_matched = TRUE) as preset_hits,
                COUNT(*) FILTER (WHERE rag_used = TRUE) as rag_queries,
                COUNT(DISTINCT user_id) as unique_users,
                COALESCE(AVG(response_time_ms)::INTEGER, 0) as avg_response_time,
                (COUNT(*) FILTER (WHERE satisfied = TRUE)::DECIMAL / 
                 NULLIF(COUNT(*) FILTER (WHERE satisfied IS NOT NULL), 0) * 100)::DECIMAL(5,2) as satisfaction_rate
            FROM hr_query_log
            WHERE created_at >= NOW() - make_interval(days => :days)
        """), {'days': days})
        
        stats = result.fetchone()
        
        total = stats[0] or 0
        preset_rate = (stats[1] / total * 100) if total > 0 else 0
        rag_rate = (stats[2] / total * 100) if total > 0 else 0
    
        return {
            'period': {
                'days': days
            },
            'stats': {
                'total_queries': total,
                'preset_hits': stats[1] or 0,
                'preset_coverage': round(preset_rate, 1),
                'rag_queries': stats[2] or 0,
                'rag_usage': round(rag_rate, 1),
                'unique_users': stats[3] or 0,
                'avg_response_time_ms': stats[4] or 0,
                'satisfaction_rate': float(stats[5]) if stats[5] else 0.0
            }
        }
    except Exception as e:
        logger.error(f"Overview error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/top-queries")
async def get_top_queries(
    days: int = 7,
    limit: int = 20,
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Get most popular queries"""
    db_session = next(get_db())
    
    try:
        result = db_session.execute(text("""
            SELECT 
                query_normalized,
                COUNT(*) as frequency,
                COUNT(*) FILTER (WHERE preset_matched = TRUE) as preset_hits,
                COUNT(*) FILTER (WHERE satisfied = TRUE) as helpful_count,
                COUNT(*) FILTER (WHERE satisfied = FALSE) as not_helpful_count,
                COUNT(*) FILTER (WHERE satisfied IS NOT NULL) as feedback_count,
                COALESCE(AVG(response_time_ms)::INTEGER, 0) as avg_response_time
            FROM hr_query_log
            WHERE created_at >= NOW() - make_interval(days => :days)
            GROUP BY query_normalized
            ORDER BY frequency DESC
            LIMIT :limit
        """), {'days': days, 'limit': limit})
        
        queries = result.fetchall()
    
        return {
            'queries': [
                {
                    'query': q[0],
                    'frequency': q[1],
                    'preset_hits': q[2],
                    'helpful_count': q[3],
                    'not_helpful_count': q[4],
                    'feedback_count': q[5],
                    'satisfaction_rate': round((q[3] / q[5] * 100) if q[5] > 0 else 0, 1),
                    'avg_response_time_ms': q[6],
                    'has_feedback': q[5] > 0
                }
                for q in queries
            ]
        }
    except Exception as e:
        logger.error(f"Top queries error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/satisfaction-breakdown")
async def get_satisfaction_breakdown(
    days: int = 7,
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Get satisfaction breakdown"""
    db_session = next(get_db())
    
    try:
        result = db_session.execute(text("""
            SELECT 
                query_normalized,
                COUNT(*) as frequency,
                COUNT(*) FILTER (WHERE satisfied = FALSE) as not_helpful
            FROM hr_query_log
            WHERE 
                created_at >= NOW() - make_interval(days => :days)
                AND satisfied = FALSE
            GROUP BY query_normalized
            ORDER BY frequency DESC
            LIMIT 10
        """), {'days': days})
        
        dissatisfied = result.fetchall()
        
        result2 = db_session.execute(text("""
            SELECT 
                query_normalized,
                COUNT(*) as frequency,
                COUNT(*) FILTER (WHERE satisfied = TRUE) as helpful
            FROM hr_query_log
            WHERE 
                created_at >= NOW() - make_interval(days => :days)
                AND satisfied = TRUE
            GROUP BY query_normalized
            ORDER BY frequency DESC
            LIMIT 10
        """), {'days': days})
        
        satisfied = result2.fetchall()
    
        return {
            'not_helpful': [
                {
                    'query': q[0],
                    'frequency': q[1],
                    'not_helpful_count': q[2]
                }
                for q in dissatisfied
            ],
            'helpful': [
                {
                    'query': q[0],
                    'frequency': q[1],
                    'helpful_count': q[2]
                }
                for q in satisfied
            ]
        }
    except Exception as e:
        logger.error(f"Satisfaction breakdown error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/preset-candidates")
async def get_preset_candidates(
    min_frequency: int = 3,
    days: int = 30,
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Find queries that should become presets"""
    db_session = next(get_db())
    
    try:
        result = db_session.execute(text("""
            SELECT 
                query_normalized,
                COUNT(*) as frequency,
                COALESCE(AVG(response_time_ms)::INTEGER, 0) as avg_time,
                COUNT(*) FILTER (WHERE satisfied = FALSE) as dissatisfied_count,
                COUNT(*) FILTER (WHERE satisfied IS NOT NULL) as feedback_count
            FROM hr_query_log
            WHERE 
                preset_matched = FALSE
                AND created_at >= NOW() - make_interval(days => :days)
            GROUP BY query_normalized
            HAVING COUNT(*) >= :min_freq
            ORDER BY frequency DESC, dissatisfied_count DESC
            LIMIT 20
        """), {'days': days, 'min_freq': min_frequency})
        
        candidates = result.fetchall()
    
        return {
            'candidates': [
                {
                    'query': c[0],
                    'frequency': c[1],
                    'avg_response_time_ms': c[2],
                    'dissatisfied_count': c[3],
                    'feedback_count': c[4],
                    'priority': 'high' if c[1] >= 10 else ('medium' if c[1] >= 5 else 'low'),
                    'should_add': c[1] >= 5 or c[3] >= 2
                }
                for c in candidates
            ]
        }
    except Exception as e:
        logger.error(f"Preset candidates error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/recent-queries")
async def get_recent_queries(
    limit: int = 50,
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Get recent queries for review"""
    db_session = next(get_db())
    
    try:
        result = db_session.execute(text("""
            SELECT 
                id,
                user_name,
                query,
                preset_matched,
                rag_used,
                response_time_ms,
                satisfied,
                created_at
            FROM hr_query_log
            ORDER BY created_at DESC
            LIMIT :limit
        """), {'limit': limit})
        
        queries = result.fetchall()
    
        return {
            'queries': [
                {
                    'id': q[0],
                    'user': q[1] or 'Anonymous',
                    'query': q[2],
                    'preset_hit': q[3],
                    'rag_used': q[4],
                    'response_ms': q[5] or 0,
                    'satisfied': q[6],
                    'timestamp': q[7].isoformat() if q[7] else None
                }
                for q in queries
            ]
        }
    except Exception as e:
        logger.error(f"Recent queries error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/hunt-roles")
async def get_hunt_roles(
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    db_session = next(get_db())
    try:
        result = db_session.execute(text("""
            SELECT DISTINCT position FROM hunt_vacancies
            WHERE position IS NOT NULL AND position != ''
            ORDER BY position
        """))
        roles = [r[0] for r in result.fetchall()]
        return {"roles": roles}
    except Exception as e:
        logger.error(f"Hunt roles error: {e}")
        return {"roles": []}


@router.get("/api/hunt-analytics")
async def get_hunt_analytics(
    role: Optional[str] = Query(default=None),
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    db_session = next(get_db())

    role_filter_v = " AND v.position ILIKE :role_pattern" if role else ""
    role_filter_c = " AND EXISTS (SELECT 1 FROM hunt_vacancies v2 WHERE v2.id = c.vacancy_id AND v2.position ILIKE :role_pattern)" if role else ""
    role_filter_p = " AND EXISTS (SELECT 1 FROM hunt_vacancies v2 WHERE v2.id = p.vacancy_id AND v2.position ILIKE :role_pattern)" if role else ""
    role_params = {"role_pattern": f"%{role}%"} if role else {}

    try:
        stats_result = db_session.execute(text(f"""
            SELECT
                (SELECT COUNT(*) FROM hunt_vacancies v WHERE 1=1{role_filter_v}) as total_vacancies,
                (SELECT COUNT(*) FROM hunt_vacancies v WHERE status = 'filled'{role_filter_v}) as total_filled,
                (SELECT COUNT(*) FROM hunt_candidates c WHERE 1=1{role_filter_c}) as total_candidates,
                (SELECT COUNT(*) FROM hunt_candidates c WHERE hr_decision = 'hired'{role_filter_c}) as total_hires,
                (SELECT COUNT(*) FROM hunt_postings p WHERE p.status = 'posted'{role_filter_p}) as total_posted,
                (SELECT COUNT(*) FROM hunt_candidates c WHERE hr_decision = 'approved'{role_filter_c}) as approved,
                (SELECT COUNT(*) FROM hunt_candidates c WHERE hr_decision = 'rejected'{role_filter_c}) as rejected,
                (SELECT COUNT(*) FROM hunt_candidates c WHERE hr_decision = 'saved'{role_filter_c}) as saved,
                (SELECT COUNT(*) FROM hunt_candidates c WHERE hr_decision = 'pending'{role_filter_c}) as pending,
                (SELECT COUNT(*) FROM hunt_candidates c WHERE created_at >= CURRENT_DATE{role_filter_c}) as candidates_today,
                (SELECT COUNT(*) FROM hunt_candidates c WHERE created_at >= CURRENT_DATE - INTERVAL '7 days'{role_filter_c}) as candidates_week,
                (SELECT COUNT(DISTINCT p.channel) FROM hunt_postings p WHERE p.status = 'posted'{role_filter_p}) as total_channels,
                (SELECT COUNT(*) FROM hunt_candidates c WHERE hr_decision NOT IN ('pending'){role_filter_c}) as reviewed
        """), role_params)
        s = stats_result.fetchone()
        total_vacancies = s[0] or 0
        total_filled = s[1] or 0
        total_candidates = s[2] or 0
        total_hires = s[3] or 0
        total_posted = s[4] or 0
        approved = s[5] or 0
        rejected = s[6] or 0
        saved = s[7] or 0
        pending = s[8] or 0
        candidates_today = s[9] or 0
        candidates_week = s[10] or 0
        total_channels = s[11] or 0
        reviewed = s[12] or 0

        shown = approved + rejected + saved + total_hires
        acceptance_rate = round((approved + total_hires) / shown * 100, 1) if shown > 0 else 0
        avg_per_vacancy = round(total_candidates / total_vacancies, 1) if total_vacancies > 0 else 0
        estimated_reach = total_posted * 5000

        from services.salary_normalizer import get_usd_uah_rate
        current_rate = get_usd_uah_rate()

        rc = ROI_CONSTANTS
        hr_rate = rc["hr_hourly_rate_uah"]
        trad_hours = rc["hours_per_vacancy_traditional"]
        portal_cost = rc["portal_cost_per_vacancy_uah"]
        maya_api_cost = rc["maya_api_cost_per_search_uah"]
        maya_minutes = rc["maya_time_per_vacancy_minutes"]

        traditional_cost_per_vacancy_uah = (hr_rate * trad_hours) + portal_cost
        maya_cost_per_vacancy_uah = maya_api_cost

        hours_saved = round(total_vacancies * (trad_hours - maya_minutes / 60), 1)
        total_savings_uah = round(total_vacancies * (traditional_cost_per_vacancy_uah - maya_cost_per_vacancy_uah))
        total_savings_usd = round(total_savings_uah / current_rate) if current_rate > 0 else 0
        cost_per_hire_maya_uah = round(maya_cost_per_vacancy_uah * total_vacancies / max(total_hires, 1))
        cost_per_hire_traditional_uah = round(traditional_cost_per_vacancy_uah)

        cost_saved_usd = total_savings_usd
        cost_per_hire_maya = round(cost_per_hire_maya_uah / current_rate, 2) if current_rate > 0 else 0

        roi_data = {
            "total_savings_uah": total_savings_uah,
            "total_savings_usd": total_savings_usd,
            "hours_saved": hours_saved,
            "cost_per_hire_maya_uah": cost_per_hire_maya_uah,
            "cost_per_hire_traditional_uah": cost_per_hire_traditional_uah,
            "cost_per_vacancy_traditional_uah": traditional_cost_per_vacancy_uah,
            "cost_per_vacancy_maya_uah": round(maya_cost_per_vacancy_uah),
            "vacancies_processed": total_vacancies,
            "total_vacancies_processed": total_vacancies,
            "total_hires": total_hires,
            "methodology": {
                "hr_hourly_rate": hr_rate,
                "hours_per_vacancy_traditional": trad_hours,
                "portal_cost": portal_cost,
                "api_cost_per_search": maya_api_cost,
                "maya_time_minutes": maya_minutes,
                "cost_per_vacancy_traditional_uah": traditional_cost_per_vacancy_uah,
                "cost_per_vacancy_maya_uah": maya_cost_per_vacancy_uah,
                "usd_rate": round(current_rate, 4),
                "traditional": {
                    "formula": "hr_rate × hours + portal_cost",
                    "hr_hourly_rate_uah": hr_rate,
                    "hours_per_vacancy": trad_hours,
                    "portal_cost_uah": portal_cost,
                    "cost_per_vacancy_uah": traditional_cost_per_vacancy_uah,
                },
                "maya": {
                    "formula": "api_cost_only",
                    "api_cost_per_search_uah": maya_api_cost,
                    "time_per_vacancy_minutes": maya_minutes,
                    "cost_per_vacancy_uah": maya_cost_per_vacancy_uah,
                },
            },
        }

        kpi = {
            "total_vacancies": total_vacancies,
            "total_filled": total_filled,
            "total_hires": total_hires,
            "total_candidates_processed": total_candidates,
            "candidates_today": candidates_today,
            "candidates_this_week": candidates_week,
            "acceptance_rate": acceptance_rate,
            "avg_candidates_per_vacancy": avg_per_vacancy,
            "total_channels_posted": total_channels,
            "estimated_reach": estimated_reach,
            "cost_saved_usd": cost_saved_usd,
            "hours_saved": hours_saved,
            "cost_per_hire_maya": cost_per_hire_maya,
            "traditional_cost_per_hire": 400,
            "decisions": {
                "approved": approved,
                "rejected": rejected,
                "saved": saved,
                "pending": pending,
                "hired": total_hires,
            },
            "roi": roi_data,
        }

        hire_funnel = {
            "found": total_candidates,
            "reviewed": reviewed,
            "approved": approved,
            "saved": saved,
            "hired": total_hires,
        }

        source_result = db_session.execute(text(f"""
            SELECT
                c.source,
                COUNT(*) as candidates_found,
                COUNT(*) FILTER (WHERE c.hr_decision = 'approved') as approved,
                COUNT(*) FILTER (WHERE c.hr_decision = 'hired') as hired
            FROM hunt_candidates c
            WHERE 1=1{role_filter_c}
            GROUP BY c.source
            ORDER BY candidates_found DESC
        """), role_params)
        source_performance = []
        for r in source_result.fetchall():
            found = r[1] or 0
            appr = r[2] or 0
            hired = r[3] or 0
            rate = round(appr / found * 100, 1) if found > 0 else 0
            source_performance.append({
                "source": r[0] or "unknown",
                "candidates_found": found,
                "approved": appr,
                "hired": hired,
                "approval_rate": rate,
            })

        posting_result = db_session.execute(text(f"""
            SELECT
                p.channel,
                COUNT(*) as total_posted,
                MAX(p.posted_at) as last_posted
            FROM hunt_postings p
            WHERE p.status = 'posted'{role_filter_p}
            GROUP BY p.channel
            ORDER BY total_posted DESC
        """), role_params)
        posting_performance = [
            {
                "channel": r[0],
                "total_posted": r[1],
                "last_posted": r[2].strftime("%Y-%m-%d %H:%M") if r[2] else None,
            }
            for r in posting_result.fetchall()
        ]

        salary_role_join = (
            " AND sd.vacancy_id IN (SELECT id FROM hunt_vacancies v WHERE v.position ILIKE :role_pattern)"
            if role else ""
        )
        salary_result = db_session.execute(text(f"""
            SELECT
                sd.city,
                sd.position,
                sd.data_type,
                COALESCE(AVG(COALESCE(sd.salary_median_usd, sd.salary_median)), 0)::INTEGER as median_usd,
                COALESCE(AVG(sd.salary_median_uah), 0)::INTEGER as median_uah,
                COUNT(*) as sample_size,
                STRING_AGG(DISTINCT sd.source, ', ') as sources
            FROM hunt_salary_data sd
            WHERE sd.city IS NOT NULL AND (sd.salary_median IS NOT NULL OR sd.salary_median_usd IS NOT NULL){salary_role_join}
            GROUP BY sd.city, sd.position, sd.data_type
            ORDER BY sd.city, sd.position
        """), role_params)
        salary_rows = salary_result.fetchall()

        city_map = {}
        for row in salary_rows:
            key = f"{row[0]}|{row[1]}"
            if key not in city_map:
                city_map[key] = {
                    "city": row[0],
                    "position": row[1],
                    "candidate_median": 0,
                    "employer_median": 0,
                    "candidate_median_uah": 0,
                    "employer_median_uah": 0,
                    "gap": 0,
                    "sample_size": 0,
                    "sources": [],
                }
            entry = city_map[key]
            if row[2] == "candidate":
                entry["candidate_median"] = row[3]
                entry["candidate_median_uah"] = row[4]
            else:
                entry["employer_median"] = row[3]
                entry["employer_median_uah"] = row[4]
            entry["sample_size"] += row[5]
            for src in (row[6] or "").split(", "):
                if src and src not in entry["sources"]:
                    entry["sources"].append(src)

        for entry in city_map.values():
            entry["gap"] = entry["candidate_median"] - entry["employer_median"]

        salary_by_city = list(city_map.values())

        skills_result = db_session.execute(text(f"""
            SELECT
                TRIM(skill) as skill,
                COUNT(*) as cnt,
                COALESCE(AVG(COALESCE(sd.salary_median_usd, sd.salary_median)), 0)::INTEGER as avg_salary_usd,
                COALESCE(AVG(sd.salary_median_uah), 0)::INTEGER as avg_salary_uah
            FROM hunt_salary_data sd,
                 LATERAL unnest(string_to_array(sd.skills, ',')) AS skill
            WHERE sd.skills IS NOT NULL AND sd.skills != ''{salary_role_join}
            GROUP BY TRIM(skill)
            ORDER BY cnt DESC
            LIMIT 10
        """), role_params)
        top_skills = [
            {"skill": r[0], "count": r[1], "avg_salary": r[2], "avg_salary_uah": r[3]}
            for r in skills_result.fetchall()
        ]

        trends_result = db_session.execute(text(f"""
            SELECT
                TO_CHAR(sd.collected_at, 'YYYY-MM') as month,
                sd.data_type,
                COALESCE(AVG(COALESCE(sd.salary_median_usd, sd.salary_median)), 0)::INTEGER as median_usd,
                COALESCE(AVG(sd.salary_median_uah), 0)::INTEGER as median_uah
            FROM hunt_salary_data sd
            WHERE (sd.salary_median IS NOT NULL OR sd.salary_median_usd IS NOT NULL){salary_role_join}
            GROUP BY TO_CHAR(sd.collected_at, 'YYYY-MM'), sd.data_type
            ORDER BY month
        """), role_params)
        trends_map = {}
        for row in trends_result.fetchall():
            if row[0] not in trends_map:
                trends_map[row[0]] = {
                    "month": row[0],
                    "candidate_median": 0, "employer_median": 0,
                    "candidate_median_uah": 0, "employer_median_uah": 0,
                }
            if row[1] == "candidate":
                trends_map[row[0]]["candidate_median"] = row[2]
                trends_map[row[0]]["candidate_median_uah"] = row[3]
            else:
                trends_map[row[0]]["employer_median"] = row[2]
                trends_map[row[0]]["employer_median_uah"] = row[3]
        salary_trends = list(trends_map.values())

        recent_result = db_session.execute(text(f"""
            SELECT
                v.id,
                v.position,
                v.city,
                v.status,
                (SELECT COUNT(*) FROM hunt_postings p WHERE p.vacancy_id = v.id AND p.status = 'posted') as channels_posted,
                (SELECT COUNT(*) FROM hunt_candidates c WHERE c.vacancy_id = v.id) as candidates_found,
                (SELECT COUNT(*) > 0 FROM hunt_candidates c WHERE c.vacancy_id = v.id AND c.hr_decision = 'hired') as has_hire,
                v.created_at
            FROM hunt_vacancies v
            WHERE 1=1{role_filter_v}
            ORDER BY v.created_at DESC
            LIMIT 10
        """), role_params)
        recent_vacancies = [
            {
                "id": r[0],
                "position": r[1] or "—",
                "city": r[2] or "—",
                "status": r[3] or "new",
                "channels_posted": r[4] or 0,
                "candidates_found": r[5] or 0,
                "hired": bool(r[6]),
                "created_at": r[7].strftime("%Y-%m-%d") if r[7] else None,
            }
            for r in recent_result.fetchall()
        ]

        source_count_result = db_session.execute(text("""
            SELECT COUNT(*) FROM hunt_sources WHERE is_active = TRUE
        """))
        active_sources = source_count_result.scalar() or 0

        mi_result = db_session.execute(text(f"""
            SELECT DISTINCT ON (sd.position, sd.data_type)
                sd.position, sd.data_type, sd.source,
                sd.salary_median_uah, sd.salary_median_usd,
                sd.salary_min_uah, sd.salary_max_uah,
                sd.usd_rate_at_collection, sd.source_url, sd.collected_at,
                COALESCE(sd.sample_count, 0) as sample_count
            FROM hunt_salary_data sd
            WHERE sd.source = 'robota.ua'{salary_role_join}
            ORDER BY sd.position, sd.data_type, sd.collected_at DESC
        """), role_params)
        mi_rows = mi_result.fetchall()
        mi_map = {}
        for r in mi_rows:
            pos = r[0]
            if pos not in mi_map:
                mi_map[pos] = {
                    "position": pos,
                    "employer_median_uah": 0, "employer_median_usd": 0,
                    "candidate_median_uah": 0, "candidate_median_usd": 0,
                    "gap_uah": 0, "gap_usd": 0,
                    "employer_count": 0, "candidate_count": 0,
                    "source": "robota.ua",
                    "source_url": r[8] or "https://robota.ua/zapros/transparent-salary",
                    "collected_at": r[9].strftime("%Y-%m-%d") if r[9] else None,
                }
            entry = mi_map[pos]
            if r[1] == "employer":
                entry["employer_median_uah"] = r[3] or 0
                entry["employer_median_usd"] = r[4] or 0
                entry["employer_count"] = r[10] or 0
            else:
                entry["candidate_median_uah"] = r[3] or 0
                entry["candidate_median_usd"] = r[4] or 0
                entry["candidate_count"] = r[10] or 0
        for entry in mi_map.values():
            entry["gap_uah"] = entry["employer_median_uah"] - entry["candidate_median_uah"]
            entry["gap_usd"] = entry["employer_median_usd"] - entry["candidate_median_usd"]
        market_intelligence = list(mi_map.values())

        return {
            "kpi": kpi,
            "hire_funnel": hire_funnel,
            "source_performance": source_performance,
            "posting_performance": posting_performance,
            "salary_by_city": salary_by_city,
            "top_skills": top_skills,
            "salary_trends": salary_trends,
            "recent_vacancies": recent_vacancies,
            "active_sources": active_sources,
            "market_intelligence": market_intelligence,
            "current_usd_rate": current_rate,
            "rate_source": "НБУ (Національний банк України)",
        }

    except Exception as e:
        logger.error(f"Hunt analytics error: {e}", exc_info=True)
        return {
            "kpi": {
                "total_vacancies": 0, "total_filled": 0, "total_hires": 0,
                "total_candidates_processed": 0, "candidates_today": 0,
                "candidates_this_week": 0, "acceptance_rate": 0,
                "avg_candidates_per_vacancy": 0, "total_channels_posted": 0,
                "estimated_reach": 0, "cost_saved_usd": 0, "hours_saved": 0,
                "cost_per_hire_maya": 0, "traditional_cost_per_hire": 400,
                "decisions": {"approved": 0, "rejected": 0, "saved": 0, "pending": 0, "hired": 0},
                "roi": {
                    "total_savings_uah": 0, "total_savings_usd": 0, "hours_saved": 0,
                    "cost_per_hire_maya_uah": 0, "cost_per_hire_traditional_uah": 0,
                    "cost_per_vacancy_traditional_uah": 0, "cost_per_vacancy_maya_uah": 0,
                    "vacancies_processed": 0, "total_hires": 0, "methodology": "",
                },
            },
            "hire_funnel": {"found": 0, "reviewed": 0, "approved": 0, "saved": 0, "hired": 0},
            "source_performance": [],
            "posting_performance": [],
            "salary_by_city": [],
            "top_skills": [],
            "salary_trends": [],
            "recent_vacancies": [],
            "active_sources": 0,
            "market_intelligence": [],
            "current_usd_rate": 41.0,
            "rate_source": "НБУ (Національний банк України)",
        }


@router.get("/api/hunt-report")
async def get_hunt_report(
    role: Optional[str] = Query(default=None),
    format: str = Query(default="json"),
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    db_session = next(get_db())

    role_filter = " AND v.position ILIKE :role_pattern" if role else ""
    role_params = {"role_pattern": f"%{role}%"} if role else {}

    try:
        from services.salary_normalizer import get_usd_uah_rate
        current_rate = get_usd_uah_rate()

        vac_result = db_session.execute(text(f"""
            SELECT
                v.id,
                v.position,
                v.city,
                v.status,
                v.salary_max,
                v.created_at,
                COUNT(c.id) as total_candidates,
                COUNT(c.id) FILTER (WHERE c.hr_decision = 'hired') as hires,
                COALESCE(AVG(c.ai_score)::INTEGER, 0) as avg_ai_score
            FROM hunt_vacancies v
            LEFT JOIN hunt_candidates c ON c.vacancy_id = v.id
            WHERE 1=1{role_filter}
            GROUP BY v.id, v.position, v.city, v.status, v.salary_max, v.created_at
            ORDER BY v.created_at DESC
        """), role_params)
        vacancies_raw = vac_result.fetchall()

        vac_ids = [r[0] for r in vacancies_raw]
        candidates_by_vac = {}
        market_by_vac = {}

        if vac_ids:
            cand_result = db_session.execute(text("""
                SELECT
                    c.vacancy_id,
                    c.id,
                    c.full_name,
                    c.age,
                    c.city,
                    c.experience_years,
                    c.salary_expectation,
                    c.ai_score,
                    c.hr_decision,
                    c.source,
                    c.profile_url,
                    c.current_role,
                    c.skills,
                    c.is_fallback,
                    c.candidate_date,
                    c.ai_summary
                FROM hunt_candidates c
                WHERE c.vacancy_id = ANY(:ids)
                ORDER BY c.ai_score DESC NULLS LAST
            """), {"ids": vac_ids})
            for r in cand_result.fetchall():
                vid = r[0]
                if vid not in candidates_by_vac:
                    candidates_by_vac[vid] = []
                salary_uah = int(r[6] * current_rate) if r[6] else None
                candidates_by_vac[vid].append({
                    "id": r[1],
                    "full_name": r[2] or "—",
                    "age": r[3],
                    "city": r[4] or "—",
                    "experience_years": r[5],
                    "salary_expectation_usd": r[6],
                    "salary_expectation_uah": salary_uah,
                    "ai_score": r[7],
                    "hr_decision": r[8] or "pending",
                    "source": r[9] or "—",
                    "profile_url": r[10],
                    "current_role": r[11],
                    "skills": r[12],
                    "is_fallback": r[13] or False,
                    "candidate_date": r[14].strftime("%Y-%m-%d") if r[14] and hasattr(r[14], 'strftime') else (str(r[14])[:10] if r[14] else None),
                    "ai_summary": r[15],
                })

            mkt_result = db_session.execute(text("""
                SELECT DISTINCT ON (sd.vacancy_id, sd.data_type)
                    sd.vacancy_id,
                    sd.data_type,
                    sd.salary_median_uah,
                    sd.salary_median_usd,
                    sd.source,
                    sd.source_url,
                    sd.salary_min_uah,
                    sd.salary_max_uah,
                    COALESCE(sd.sample_count, 0),
                    sd.collected_at
                FROM hunt_salary_data sd
                WHERE sd.vacancy_id = ANY(:ids)
                ORDER BY sd.vacancy_id, sd.data_type, sd.collected_at DESC
            """), {"ids": vac_ids})
            for r in mkt_result.fetchall():
                vid = r[0]
                if vid not in market_by_vac:
                    market_by_vac[vid] = {"candidate": {}, "employer": {}}
                dtype = r[1] or "employer"
                market_by_vac[vid][dtype] = {
                    "salary_median_uah": r[2],
                    "salary_median_usd": r[3],
                    "source": r[4],
                    "source_url": r[5],
                    "salary_min_uah": r[6],
                    "salary_max_uah": r[7],
                    "sample_count": r[8],
                    "collected_at": r[9].strftime("%Y-%m-%d") if r[9] and hasattr(r[9], 'strftime') else (str(r[9])[:10] if r[9] else None),
                }

        vacancies_json = []
        for r in vacancies_raw:
            vid = r[0]
            mkt = market_by_vac.get(vid, {})
            mkt_median_uah = (mkt.get("employer") or mkt.get("candidate") or {}).get("salary_median_uah")
            mkt_source = (mkt.get("employer") or mkt.get("candidate") or {}).get("source", "—")
            salary_max_uah = int(r[4] * current_rate) if r[4] else None
            mkt_source_url = (mkt.get("employer") or mkt.get("candidate") or {}).get("source_url")
            vacancies_json.append({
                "id": vid,
                "position": r[1] or "—",
                "city": r[2] or "—",
                "status": r[3] or "new",
                "salary_max_usd": r[4],
                "salary_max_uah": salary_max_uah,
                "created_at": r[5].strftime("%Y-%m-%d") if r[5] else None,
                "total_candidates": r[6] or 0,
                "hires": r[7] or 0,
                "avg_ai_score": r[8] or 0,
                "market_median_uah": mkt_median_uah,
                "market_source": mkt_source,
                "market_source_url": mkt_source_url,
                "market_data": mkt,
                "candidates": candidates_by_vac.get(vid, []),
            })

        total_hires_report = sum(v["hires"] for v in vacancies_json)
        total_candidates_report = sum(v["total_candidates"] for v in vacancies_json)
        avg_score = (
            sum(v["avg_ai_score"] for v in vacancies_json if v["avg_ai_score"]) /
            max(sum(1 for v in vacancies_json if v["avg_ai_score"]), 1)
        )

        hire_rate = round(total_hires_report / total_candidates_report * 100, 1) if total_candidates_report > 0 else 0
        if format == "json":
            return {
                "vacancies": vacancies_json,
                "current_usd_rate": current_rate,
                "total_vacancies": len(vacancies_json),
                "role_filter": role,
                "summary": {
                    "total_vacancies": len(vacancies_json),
                    "total_candidates": total_candidates_report,
                    "total_hires": total_hires_report,
                    "hire_rate": hire_rate,
                    "avg_ai_score": round(avg_score, 1),
                    "avg_time_to_fill_hours": None,
                    "current_usd_rate": current_rate,
                },
            }

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter

        BLUE_FILL = PatternFill("solid", fgColor="4472C4")
        LIGHT_FILL = PatternFill("solid", fgColor="F2F2F2")
        HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        BOLD_FONT = Font(name="Calibri", bold=True, size=10)
        NORMAL_FONT = Font(name="Calibri", size=10)
        CURRENCY_FORMAT = '#,##0'

        def style_header_row(ws, cols):
            for col_idx, col_name in enumerate(cols, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = HEADER_FONT
                cell.fill = BLUE_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def auto_fit(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

        STATUS_UK = {
            "searching": "Пошук", "completed": "Завершено", "posted": "Опубліковано",
            "filled": "Закрито", "new": "Новий", "pending": "Очікування", "no_results": "Без результатів",
        }
        DECISION_UK = {
            "hired": "Найнято", "approved": "Схвалено", "rejected": "Відхилено",
            "saved": "Збережено", "pending": "Очікування",
        }
        DTYPE_UK = {"employer": "Роботодавець", "candidate": "Кандидат"}

        def write_row(ws, row_idx, values, fill, currency_cols=(), hyperlink_col_val=None):
            for col_idx, val in enumerate(values, start=1):
                if hyperlink_col_val and col_idx == hyperlink_col_val[0] and hyperlink_col_val[1]:
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.hyperlink = hyperlink_col_val[1]
                    cell.font = Font(name="Calibri", size=10, color="1E40AF", underline="single")
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = NORMAL_FONT
                if fill:
                    cell.fill = fill
                if col_idx in currency_cols:
                    cell.number_format = CURRENCY_FORMAT

        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Звіт по вакансіях"
        vac_cols = [
            "ID", "Посада", "Місто", "Статус",
            "ЗП макс (USD)", "ЗП макс (UAH)",
            "Кандидатів", "Avg AI оцінка", "Найнято",
            "Медіана роботодавця (UAH)", "Медіана кандидата (UAH)", "Розрив ЗП (UAH)",
            "Джерело ринку", "Дата створення",
        ]
        style_header_row(ws1, vac_cols)
        ws1.freeze_panes = "A2"
        for row_idx, v in enumerate(vacancies_json, start=2):
            fill = LIGHT_FILL if row_idx % 2 == 0 else None
            mkt_src_url = v.get("market_source_url") or ""
            mkt_src_label = v.get("market_source") or "—"
            mkt = v.get("market_data", {})
            emp_median = (mkt.get("employer") or {}).get("salary_median_uah")
            cand_median = (mkt.get("candidate") or {}).get("salary_median_uah")
            gap_uah = (cand_median - emp_median) if emp_median and cand_median else None
            row_data = [
                v["id"], v["position"], v["city"],
                STATUS_UK.get(v["status"], v["status"]),
                v["salary_max_usd"], v["salary_max_uah"],
                v["total_candidates"], v["avg_ai_score"], v["hires"],
                emp_median, cand_median, gap_uah,
                mkt_src_label, v["created_at"],
            ]
            write_row(ws1, row_idx, row_data, fill,
                      currency_cols=(5, 6, 10, 11, 12),
                      hyperlink_col_val=(13, mkt_src_url) if mkt_src_url else None)
        auto_fit(ws1)

        ws2 = wb.create_sheet("Кандидати")
        cand_cols = [
            "Вакансія ID", "Посада", "ПІБ", "Вік", "Місто", "Досвід (років)",
            "Поточна посада", "Навички", "ЗП очік. (USD)", "ЗП очік. (UAH)",
            "AI оцінка", "Резюме AI", "Рішення", "Джерело", "Профіль (URL)",
            "Дата кандидата", "Fallback",
        ]
        style_header_row(ws2, cand_cols)
        ws2.freeze_panes = "A2"
        cand_row = 2
        for v in vacancies_json:
            for c in v["candidates"]:
                fill = LIGHT_FILL if cand_row % 2 == 0 else None
                profile_url = c.get("profile_url") or ""
                row_data = [
                    v["id"], v["position"], c["full_name"], c["age"],
                    c["city"], c["experience_years"],
                    c.get("current_role"), c.get("skills"),
                    c["salary_expectation_usd"], c["salary_expectation_uah"],
                    c["ai_score"], c.get("ai_summary"),
                    DECISION_UK.get(c["hr_decision"], c["hr_decision"]),
                    c["source"], profile_url,
                    c.get("candidate_date"),
                    "Так" if c.get("is_fallback") else "Ні",
                ]
                write_row(ws2, cand_row, row_data, fill,
                          currency_cols=(9, 10),
                          hyperlink_col_val=(15, profile_url) if profile_url else None)
                cand_row += 1
        auto_fit(ws2)

        ws3 = wb.create_sheet("Аналітика ринку")
        mkt_cols = [
            "Посада", "Місто",
            "Медіана роботодавця (UAH)", "Медіана роботодавця (USD)",
            "Медіана кандидата (UAH)", "Медіана кандидата (USD)",
            "Розрив (UAH)", "Розрив (USD)",
            "Джерело", "URL джерела",
        ]
        style_header_row(ws3, mkt_cols)
        ws3.freeze_panes = "A2"
        mkt_row = 2
        for v in vacancies_json:
            mkt = market_by_vac.get(v["id"], {})
            emp = mkt.get("employer") or {}
            cnd = mkt.get("candidate") or {}
            emp_med_uah = emp.get("salary_median_uah")
            emp_med_usd = emp.get("salary_median_usd")
            cnd_med_uah = cnd.get("salary_median_uah")
            cnd_med_usd = cnd.get("salary_median_usd")
            gap_uah = (cnd_med_uah - emp_med_uah) if emp_med_uah and cnd_med_uah else None
            gap_usd = (cnd_med_usd - emp_med_usd) if emp_med_usd and cnd_med_usd else None
            src_url = emp.get("source_url") or cnd.get("source_url") or ""
            src_label = emp.get("source") or cnd.get("source") or "—"
            fill = LIGHT_FILL if mkt_row % 2 == 0 else None
            row_data = [
                v["position"], v["city"],
                emp_med_uah, emp_med_usd,
                cnd_med_uah, cnd_med_usd,
                gap_uah, gap_usd,
                src_label, src_url,
            ]
            write_row(ws3, mkt_row, row_data, fill,
                      currency_cols=(3, 4, 5, 6, 7, 8),
                      hyperlink_col_val=(10, src_url) if src_url else None)
            mkt_row += 1
        auto_fit(ws3)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = "hunt_report.xlsx"
        if role:
            safe_role = role.replace(" ", "_")[:30]
            filename = f"hunt_report_{safe_role}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
        )

    except Exception as e:
        logger.error(f"Hunt report error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/pulse-overview")
async def get_pulse_overview(
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Get Team Pulse statistics: mood trends, department breakdown, trigger counts."""
    db_session = next(get_db())
    try:
        monthly_mood = db_session.execute(text("""
            SELECT
                survey_month,
                ROUND(AVG(score)::NUMERIC, 2) AS avg_score,
                COUNT(*) AS response_count
            FROM pulse_surveys
            WHERE survey_month >= TO_CHAR(NOW() - INTERVAL '6 months', 'YYYY-MM')
            GROUP BY survey_month
            ORDER BY survey_month
        """)).fetchall()

        dept_mood = db_session.execute(text("""
            SELECT
                COALESCE(department, 'Невідомо') AS department,
                ROUND(AVG(score)::NUMERIC, 2) AS avg_score,
                COUNT(*) AS response_count
            FROM pulse_surveys
            WHERE survey_month = TO_CHAR(NOW(), 'YYYY-MM')
            GROUP BY department
            ORDER BY avg_score DESC
        """)).fetchall()

        trigger_counts = db_session.execute(text("""
            SELECT
                trigger_type,
                COUNT(*) AS cnt
            FROM pulse_triggers
            WHERE fired_at >= DATE_TRUNC('month', NOW())
            GROUP BY trigger_type
            ORDER BY cnt DESC
        """)).fetchall()

        total_users_row = db_session.execute(text(
            "SELECT COUNT(*) FROM hr_users WHERE is_active = TRUE AND telegram_id IS NOT NULL"
        )).scalar() or 0

        current_month = db_session.execute(text(
            "SELECT COUNT(*) FROM pulse_surveys "
            "WHERE survey_month = TO_CHAR(NOW(), 'YYYY-MM')"
        )).scalar() or 0

        overall_avg = db_session.execute(text(
            "SELECT ROUND(AVG(score)::NUMERIC, 2) FROM pulse_surveys "
            "WHERE survey_month = TO_CHAR(NOW(), 'YYYY-MM')"
        )).scalar()

        response_rate = round((current_month / total_users_row * 100), 1) if total_users_row > 0 else 0.0

        monthly_by_dept_rows = db_session.execute(text("""
            SELECT
                survey_month,
                COALESCE(department, 'Невідомо') AS department,
                ROUND(AVG(score)::NUMERIC, 2) AS avg_score,
                COUNT(*) AS response_count
            FROM pulse_surveys
            WHERE survey_month >= TO_CHAR(NOW() - INTERVAL '6 months', 'YYYY-MM')
            GROUP BY survey_month, department
            ORDER BY survey_month, department
        """)).fetchall()

        individual_responses = db_session.execute(text("""
            SELECT
                COALESCE(employee_name, user_hash, 'Анонім') AS name,
                COALESCE(department, 'Невідомо') AS department,
                score,
                COALESCE(problem_category, '') AS problem_category,
                COALESCE(problem_text, '') AS problem_text,
                COALESCE(responded_at::text, '') AS responded_at
            FROM pulse_surveys
            WHERE survey_month = TO_CHAR(NOW(), 'YYYY-MM')
            ORDER BY score ASC, responded_at DESC
        """)).fetchall()

        problem_breakdown = db_session.execute(text("""
            SELECT
                COALESCE(problem_category, 'none') AS category,
                COUNT(*) AS cnt
            FROM pulse_surveys
            WHERE survey_month = TO_CHAR(NOW(), 'YYYY-MM')
              AND score = 1
            GROUP BY problem_category
            ORDER BY cnt DESC
        """)).fetchall()

        return {
            "monthly_mood": [
                {"month": r[0], "avg_score": float(r[1] or 0), "responses": r[2]}
                for r in monthly_mood
            ],
            "dept_mood": [
                {"department": r[0], "avg_score": float(r[1] or 0), "responses": r[2]}
                for r in dept_mood
            ],
            "monthly_by_department": [
                {
                    "month": r[0],
                    "department": r[1],
                    "avg_score": float(r[2] or 0),
                    "responses": r[3],
                }
                for r in monthly_by_dept_rows
            ],
            "trigger_counts": [
                {"trigger_type": r[0], "count": r[1]}
                for r in trigger_counts
            ],
            "individual_responses": [
                {
                    "name": r[0],
                    "department": r[1],
                    "score": r[2],
                    "problem_category": r[3],
                    "problem_text": r[4],
                    "responded_at": r[5],
                }
                for r in individual_responses
            ],
            "problem_breakdown": [
                {"category": r[0], "count": r[1]}
                for r in problem_breakdown
            ],
            "kpi": {
                "response_rate": response_rate,
                "responses_this_month": int(current_month),
                "total_active_users": int(total_users_row),
                "overall_mood_avg": float(overall_avg) if overall_avg else 0.0,
                "triggers_this_month": sum(r[1] for r in trigger_counts),
            }
        }
    except Exception as e:
        logger.error(f"Pulse overview error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db_session.close()


@router.get("/api/pulse-alerts")
async def get_pulse_alerts(
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Return employees with elevated risk scores (current_score >= 4)."""
    db_session = next(get_db())
    try:
        rows = db_session.execute(text("""
            SELECT
                prs.employee_id,
                prs.employee_name,
                prs.department,
                prs.current_score,
                prs.alert_status,
                prs.last_trigger_at,
                pt.trigger_type AS last_trigger_type,
                pt.trigger_text AS last_trigger_text,
                pt.id AS last_trigger_id
            FROM pulse_risk_scores prs
            LEFT JOIN pulse_triggers pt
                ON pt.employee_id = prs.employee_id
                AND pt.id = (
                    SELECT id FROM pulse_triggers
                    WHERE employee_id = prs.employee_id
                    ORDER BY fired_at DESC LIMIT 1
                )
            WHERE prs.current_score >= 4
            ORDER BY prs.current_score DESC, prs.last_trigger_at DESC
        """)).fetchall()

        return {
            "alerts": [
                {
                    "employee_id": r[0],
                    "employee_name": r[1] or "Невідомий",
                    "department": r[2] or "Невідомий відділ",
                    "current_score": r[3],
                    "alert_status": r[4] or "red",
                    "last_trigger_at": r[5].isoformat() if r[5] else None,
                    "last_trigger_type": r[6],
                    "last_trigger_text": r[7],
                    "last_trigger_id": r[8],
                }
                for r in rows
            ],
            "total": len(rows),
        }
    except Exception as e:
        logger.error(f"Pulse alerts error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db_session.close()


@router.get("/api/pulse-videos")
async def get_pulse_video_stats(
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Return video view counts for pulse support videos."""
    db_session = next(get_db())
    try:
        rows = db_session.execute(text("""
            SELECT
                video_id,
                COUNT(*) AS view_count,
                COUNT(CASE WHEN completed THEN 1 END) AS completed_count,
                MAX(viewed_at) AS last_viewed_at
            FROM pulse_video_views
            WHERE viewed_at >= NOW() - INTERVAL '30 days'
            GROUP BY video_id
            ORDER BY view_count DESC
        """)).fetchall()

        return {
            "videos": [
                {
                    "video_id": r[0],
                    "view_count": r[1],
                    "completed_count": r[2],
                    "last_viewed_at": r[3].isoformat() if r[3] else None,
                }
                for r in rows
            ]
        }
    except Exception as e:
        logger.error(f"Pulse videos error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db_session.close()


@router.get("/api/pulse-risk-history/{employee_id}")
async def get_pulse_risk_history(
    employee_id: int,
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Return recent trigger history for a specific employee."""
    from services.pulse_service import get_risk_history as _get_risk_history
    return {"history": _get_risk_history(employee_id)}


@router.post("/api/pulse-hr-action")
async def post_pulse_hr_action(
    request: Request,
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Record an HR action on a trigger (resolved / false_positive / escalated)."""
    body = await request.json()
    trigger_id = body.get("trigger_id")
    action = body.get("action")
    hr_user = body.get("hr_user", "admin")

    if not trigger_id or not action:
        raise HTTPException(status_code=400, detail="trigger_id and action required")

    from services.pulse_service import log_hr_action as _log_hr_action
    _log_hr_action(trigger_id, action, hr_user)
    return {"ok": True}


@router.get("/api/system-info")
async def get_system_info(
    credentials: HTTPBasicCredentials = Depends(verify_admin)
):
    """Get system information"""
    db_session = next(get_db())
    
    try:
        content_result = db_session.execute(text("SELECT COUNT(*) FROM hr_content"))
        content_count = content_result.scalar() or 0
        
        try:
            preset_result = db_session.execute(text(
                "SELECT COUNT(*) FROM hr_preset_answers WHERE is_active = TRUE"
            ))
            preset_count = preset_result.scalar() or 0
        except:
            preset_count = 6
        
        try:
            embedding_result = db_session.execute(text("SELECT COUNT(*) FROM hr_embeddings"))
            embedding_count = embedding_result.scalar() or 0
        except:
            embedding_count = 0
    
        return {
            'content_items': content_count,
            'active_presets': preset_count,
            'embeddings': embedding_count,
            'status': 'operational'
        }
    except Exception as e:
        logger.error(f"System info error: {e}")
        return {
            'content_items': 0,
            'active_presets': 6,
            'embeddings': 0,
            'status': 'operational'
        }


@router.get("/api/pulse-test-survey")
async def test_pulse_survey(
    tid: int = None,
    credentials: HTTPBasicCredentials = Depends(verify_admin),
):
    """Send survey to one user (by telegram_id) or all active users if no tid given."""
    from services.pulse_service import send_monthly_survey, send_survey_to_user
    if tid:
        count = await send_survey_to_user(tid)
    else:
        count = send_monthly_survey()
    return {"sent_to": count}


@router.get("/api/survey/{survey_id}/results")
async def get_survey_results(
    survey_id: str,
    credentials: HTTPBasicCredentials = Depends(verify_admin),
):
    """Return current vote counts and percentages for a survey."""
    from services.survey_service import get_results
    return await get_results(survey_id)


@router.post("/api/survey/{survey_id}/close")
async def close_survey_endpoint(
    survey_id: str,
    credentials: HTTPBasicCredentials = Depends(verify_admin),
):
    """Manually close a survey and post final scoreboard to observers."""
    from services.survey_service import close_survey
    await close_survey(survey_id)
    return {"status": "closed", "survey_id": survey_id}


